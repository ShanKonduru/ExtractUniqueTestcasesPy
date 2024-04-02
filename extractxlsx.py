import openpyxl

def extract_test_case_details(input_file, output_file):
    # Load the input Excel workbook
    wb_in = openpyxl.load_workbook(input_file)
    ws_test_cases = wb_in['Train_Regression_TestCases']
    ws_test_names = wb_in['Unique names']

    # Read unique test case names from "Unique names" sheet, skipping the first row
    test_names = [cell.value for cell in ws_test_names['A'][1:] if cell.value]

    # Create a dictionary to store rows associated with each test case name
    test_case_rows = {test_name: [] for test_name in test_names}

    # Initialize variables to track current test case name and its rows
    current_test_case_name = None
    current_test_case_rows = []

    # Iterate through the "Train_Regression_TestCases" sheet to extract rows
    for row in ws_test_cases.iter_rows(min_row=2, values_only=True):
        if row[2] in test_names:  # Check if current row has a test case name
            # If we already collected rows for a previous test case, store them
            if current_test_case_name is not None:
                test_case_rows[current_test_case_name] = current_test_case_rows
            # Start collecting rows for the new test case
            current_test_case_name = row[2]
            current_test_case_rows = [row]
        else:
            # Collect rows for the current test case until a new test case name is encountered
            current_test_case_rows.append(row)

    # Store the rows for the last test case encountered
    if current_test_case_name is not None:
        test_case_rows[current_test_case_name] = current_test_case_rows

    # Create a new Excel workbook and sheet for the output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    # Copy header from input to output
    for col_num, cell in enumerate(ws_test_cases[1], 1):
        ws_out.cell(row=1, column=col_num, value=cell.value)

    # Write rows to output sheet maintaining the order of test case names
    current_row = 2  # Start from row 2 as row 1 contains headers
    for test_name in test_names:
        rows = test_case_rows[test_name]
        for row in rows:
            for col_num, value in enumerate(row, 1):
                ws_out.cell(row=current_row, column=col_num, value=value)
            current_row += 1

    # Save the output Excel workbook
    wb_out.save(output_file)

if __name__ == '__main__':
    input_file = './Testdata/Train_Regression_Testcases.xlsx'
    output_file = './Output/output_Train_Regression_Testcases.xlsx'
    
    extract_test_case_details(input_file, output_file)
    print(f"Test case details extracted and saved to {output_file}.")

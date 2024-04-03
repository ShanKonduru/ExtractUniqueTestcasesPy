import openpyxl

class ExcelTestCaseExtractor:
    def __init__(self, input_file, testCases_Data, unique_names, output_file):
        self.testCasesData = testCases_Data
        self.uniquenames = unique_names
        self.input_file = input_file
        self.output_file = output_file
        self.wb_in = openpyxl.load_workbook(self.input_file)
        self.ws_test_cases = self.wb_in[self.testCasesData]
        self.ws_test_names = self.wb_in[self.uniquenames]
        self.test_names = self._get_test_names()
        self.test_case_rows = self._extract_test_case_rows()

    def _get_test_names(self):
        return [cell.value for cell in self.ws_test_names['A'][1:] if cell.value]

    def _extract_test_case_rows(self):
        test_case_rows = {test_name: [] for test_name in self.test_names}
        current_test_case_name = None
        current_test_case_rows = []

        for row in self.ws_test_cases.iter_rows(min_row=2, values_only=True):
            if row[2] in self.test_names:
                if current_test_case_name is not None:
                    test_case_rows[current_test_case_name] = current_test_case_rows
                current_test_case_name = row[2]
                current_test_case_rows = [row]
            else:
                current_test_case_rows.append(row)

        if current_test_case_name is not None:
            test_case_rows[current_test_case_name] = current_test_case_rows

        return test_case_rows

    def _copy_header(self, ws_from, ws_to):
        for col_num, cell in enumerate(ws_from[1], 1):
            ws_to.cell(row=1, column=col_num, value=cell.value)

    def _write_rows_to_sheet(self, ws, rows, start_row):
        current_row = start_row
        for row in rows:
            for col_num, value in enumerate(row, 1):
                ws.cell(row=current_row, column=col_num, value=value)
            current_row += 1

    def save_output(self):
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active

        self._copy_header(self.ws_test_cases, ws_out)

        current_row = 2
        for test_name in self.test_names:
            rows = self.test_case_rows[test_name]
            self._write_rows_to_sheet(ws_out, rows, current_row)
            current_row += len(rows)

        wb_out.save(self.output_file)
        print(f"Test case details extracted and saved to {self.output_file}.")

if __name__ == '__main__':
    input_file = './Testdata/EME_Regression_Test Cases.xlsx'
    output_file = './Output/output_EME_Regression_Test Cases.xlsx'

    extractor = ExcelTestCaseExtractor(input_file, 'TestCases Data', 'Unique names', output_file)
    extractor.save_output()

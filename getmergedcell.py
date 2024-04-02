import openpyxl

def copy_merged_cell_data_to_new_sheet(file_path, source_sheet_name, target_sheet_name, column_identifier):
    """
    Copy merged cell data and their associated rows to a new sheet in the same Excel workbook.
    
    Args:
    - file_path (str): Path to the Excel file.
    - source_sheet_name (str): Name of the source worksheet.
    - target_sheet_name (str): Name of the target worksheet to copy data to.
    - column_identifier (str or int): Column Name (e.g., 'A', 'B', 'C') or Column Number (e.g., 1, 2, 3) of the cell.
    """
    
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the source worksheet
    ws_source = wb[source_sheet_name]
    
    # Convert column identifier to column index if it's a column name
    if isinstance(column_identifier, str):
        col_index = openpyxl.utils.column_index_from_string(column_identifier)
    else:
        col_index = column_identifier
    
    # Initialize variables
    merged_cell_data_list = []
    visited_cells = set()
    
    # Check each row in the specified column for merged cells
    for row in range(1, ws_source.max_row + 1):
        cell = ws_source.cell(row=row, column=col_index)
        if cell.coordinate in ws_source.merged_cells and cell.coordinate not in visited_cells:
            # Get the dimensions of the merged cell
            for merged_range in ws_source.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
                    num_rows = max_row - min_row + 1
                    num_columns = max_col - min_col + 1
                    
                    # Get row data associated with the merged cell
                    row_data = [ws_source.cell(row=row_idx, column=col_idx).value for row_idx in range(min_row, max_row + 1) for col_idx in range(min_col, max_col + 1)]
                    
                    merged_cell_data_list.append((cell.coordinate, num_rows, num_columns, row_data))
                    visited_cells.update(merged_range)
                    break
    
    # Create or select the target worksheet
    if target_sheet_name in wb.sheetnames:
        ws_target = wb[target_sheet_name]
    else:
        ws_target = wb.create_sheet(title=target_sheet_name)
    
    # Write merged cell data to target sheet
    for idx, (_, _, _, row_data) in enumerate(merged_cell_data_list, start=1):
        # Write merged cell data to target sheet
        for col_idx, value in enumerate(row_data, start=1):
            ws_target.cell(row=idx, column=col_idx, value=value)

    # Save the workbook
    wb.save(file_path)
    
    print(f"Merged cell data copied to {target_sheet_name} sheet.")

# Input parameters
file_path = './SampleFile.xlsx'
source_sheet_name = 'TestData'
target_sheet_name = 'CopiedData'
column_identifier = 'A'  # Can also be an integer (e.g., 1 for column A)

# Copy merged cell data to new sheet
copy_merged_cell_data_to_new_sheet(file_path, source_sheet_name, target_sheet_name, column_identifier)
